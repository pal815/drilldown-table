# -*- coding: utf-8 -*-
"""입력 '표 영역만' 가이드 이미지 2종 생성(제목/단위/코너 없이 = 채워야 할 표만).
   → 사용자가 '표만 채우면 되는지' 헷갈리지 않게. assets/previews/guide_*.png"""
import os
import sys
import time
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image
import win32com.client as win32

HEAD = PatternFill("solid", fgColor="D6E0F2")
BOLD = Font(bold=True)
CEN = Alignment(horizontal="center")
_t = Side(style="thin", color="BFBFBF")
BORD = Border(left=_t, right=_t, top=_t, bottom=_t)
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "previews")
RES = getattr(Image, "Resampling", Image).LANCZOS


def build_rc(ws):
    for j, h in enumerate(["항목", "세부", "매출", "영업이익", "점유율"], 1):
        c = ws.cell(1, j, h); c.fill = HEAD; c.font = BOLD; c.alignment = CEN; c.border = BORD
    rows = [("A사업부", "계", 1200, 180, "21%"), ("", "Q1", 280, 38, "20%"),
            ("", "Q2", 300, 42, "21%"), ("", "Q3", 300, 48, "22%"), ("", "Q4", 320, 52, "21%"),
            ("B사업부", "계", 800, 90, "12%"), ("", "Q1", 180, 18, "11%"),
            ("", "Q2", 200, 22, "12%"), ("", "Q3", 210, 24, "12%"), ("", "Q4", 210, 26, "13%")]
    for i, row in enumerate(rows, 2):
        for j, v in enumerate(row, 1):
            c = ws.cell(i, j, v); c.border = BORD
    for col, w in zip("ABCDE", (12, 8, 9, 11, 9)):
        ws.column_dimensions[col].width = w


def build_both(ws):
    a = ws.cell(2, 1, "행항목"); b = ws.cell(2, 2, "행세부")
    for c in (a, b):
        c.fill = HEAD; c.font = BOLD; c.border = BORD
    for yr, c0 in (("2023", 3), ("2024", 5)):
        ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c0 + 1)
        yc = ws.cell(1, c0, yr); yc.fill = HEAD; yc.font = BOLD; yc.alignment = CEN
        for off, half in enumerate(("상반기", "하반기")):
            hc = ws.cell(2, c0 + off, half); hc.fill = HEAD; hc.font = BOLD; hc.alignment = CEN; hc.border = BORD
    data = [("A사업부", "국내", 700, 750, 780, 820), ("", "해외", 450, 500, 470, 530),
            ("B사업부", "국내", 300, 320, 310, 330), ("", "해외", 250, 260, 270, 280)]
    for i, row in enumerate(data, 3):
        for j, v in enumerate(row, 1):
            c = ws.cell(i, j, v); c.border = BORD
    for col, w in zip("ABCDEF", (12, 9, 9, 9, 9, 9)):
        ws.column_dimensions[col].width = w


jobs = []
for name, builder in (("guide_rowcol", build_rc), ("guide_both", build_both)):
    wb = openpyxl.Workbook(); builder(wb.active)
    tmp = os.path.join(tempfile.gettempdir(), name + ".xlsx"); wb.save(tmp)
    jobs.append((name, tmp))

excel = win32.Dispatch("Excel.Application")
excel.Visible = False; excel.DisplayAlerts = False
try:
    for name, tmp in jobs:
        wb = excel.Workbooks.Open(tmp); ws = wb.Worksheets(1); ws.Activate()
        rng = ws.UsedRange
        for _ in range(4):
            try:
                rng.CopyPicture(1, 2); break
            except Exception:
                time.sleep(0.6)
        time.sleep(0.4)
        co = ws.ChartObjects().Add(0, 0, rng.Width, rng.Height); ch = co.Chart
        ch.ChartArea.Format.Fill.Visible = False; ch.ChartArea.Format.Line.Visible = False
        co.Activate(); time.sleep(0.2); ch.Paste(); time.sleep(0.2)
        png = os.path.join(OUT, name + ".png"); ch.Export(png, "PNG"); co.Delete()
        wb.Close(False)
        im = Image.open(png)
        if im.width > 760:
            im = im.resize((760, int(im.height * 760 / im.width)), RES)
        im.save(png); print("guide:", name, im.size)
finally:
    excel.Quit()
