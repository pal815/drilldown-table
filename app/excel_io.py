# -*- coding: utf-8 -*-
"""
excel_io.py — 사용자용 '입력양식.xlsx' <-> drilldown_table 엔진 모델(dict) 변환기.

엔진(drilldown_table.py)의 양식 로직은 일절 건드리지 않는다. 이 모듈은
(1) 비개발자가 채우는 입력양식 엑셀을 만들고(make_template),
(2) 채워진 엑셀을 엔진이 먹는 model dict 으로 읽어온다(read_model).
YAML/JSON 을 사용자에게 노출하지 않기 위한 '엑셀 입력 어댑터'.

레이아웃 규약
─ 행·열 시트(orient=row/column 공용) ─
  A1=제목  B1=<값>   A2=단위 B2=<값>   A3=코너라벨 B3=<값>
  5행 헤더: A5=항목  B5=세부  C5~=속성(매출/영업이익/점유율 …)
  6행~ 데이터: A=항목(그룹 첫 행에만), B=세부.  B가 '계/합계'면 그 항목의 합계줄,
              그 외면 세부 행. 값은 C열부터 속성 수만큼.
─ 행+열(교차) 시트(orient=both) ─
  A1=제목 / A2=단위 / A3=코너라벨 (B열에 값)
  5행: C5~=열그룹 라벨(예: 2023, 2024)  6행: C6~=열세부(상반기/하반기)
  7행~: A=행그룹, B=행세부, C~=교차 값.  계/총계는 엔진이 자동 합산.
"""
import os
import re
import tempfile

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SHEET_ROWCOL = "행·열"
SHEET_BOTH = "행+열(교차)"
_SUMMARY_LABELS = {"계", "합계", "소계", "합", "total", "sum"}


# ── 셀 값 읽기 (퍼센트 서식 → "NN%" 문자열) ──────────────────────────────
def _val(cell):
    v = cell.value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    fmt = cell.number_format or ""
    if "%" in fmt:  # Excel은 21% 를 0.21+퍼센트서식으로 저장 → "21%" 문자열 복원
        pct = v * 100
        return f"{int(round(pct))}%" if abs(pct - round(pct)) < 1e-9 else f"{pct:g}%"
    return v


def _s(v):
    return str(v).strip() if v not in (None, "") else ""


# ── 읽기: 행·열 시트 → model (depth 자동감지: 2단계는 items, 3+단계는 nodes) ──
def _is_num_token(v):
    """_val() 산출물 기준 숫자 판정(퍼센트 "21%"·천단위·부호 포함). 빈칸/라벨=False."""
    if v is None or isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return True
    s = str(v).strip().replace(",", "")
    if s.endswith("%"):
        s = s[:-1]
    s = s.lstrip("+-")
    if s == "":
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def _detect_label_cols(ws, header_row=5, data_start=6):
    """'숫자만 시작하는 열'의 왼쪽 라벨 열 개수 = depth(가장 깊은 가지 기준)."""
    W = ws.max_column or 1
    maxlab = 0
    for r in range(data_start, ws.max_row + 1):
        vals = [_val(ws.cell(r, c)) for c in range(1, W + 1)]
        if not any(_is_num_token(x) for x in vals):
            continue                                   # 값 없는 행은 depth 추정에서 제외
        last_label = 0
        for c in range(1, W + 1):
            x = vals[c - 1]
            if x is not None and not _is_num_token(x):
                last_label = c                         # 텍스트 = 라벨(빈칸 None은 경계 미갱신)
        maxlab = max(maxlab, last_label)
    return max(maxlab, 1)


def _node_vals(n):
    return n["summary"] if n.get("children") is not None else n.get("values")


def _autosum(nodes, A):
    """summary 없는 그룹: 숫자 칸=자식 합, 퍼센트/문자 칸=공백."""
    for n in nodes:
        if n.get("children") is not None:
            _autosum(n["children"], A)
            cur = n.get("summary")
            if cur is None or all(v is None for v in cur):
                acc, seen, ok = [0] * A, [False] * A, [True] * A
                for ch in n["children"]:
                    cv = _node_vals(ch) or [None] * A
                    for j in range(A):
                        x = cv[j] if j < len(cv) else None
                        if isinstance(x, (int, float)) and not isinstance(x, bool):
                            acc[j] += x
                            seen[j] = True
                        elif x is not None:
                            ok[j] = False              # 퍼센트/문자 → 합산 불가
                n["summary"] = [acc[j] if (ok[j] and seen[j]) else None for j in range(A)]


def _read_rowcol_ndepth(ws, LABELCOLS, header_row=5, data_start=6):
    attrs, c = [], LABELCOLS + 1
    while _s(ws.cell(header_row, c).value):
        attrs.append(_s(ws.cell(header_row, c).value))
        c += 1
    A = len(attrs)
    if A == 0:
        raise ValueError("값(숫자) 열이 없습니다 — 라벨 오른쪽에 속성 헤더(매출 등)를 적어주세요.")
    level_labels = [_s(ws.cell(header_row, c).value) for c in range(1, LABELCOLS + 1)]

    roots, path = [], [None] * LABELCOLS    # path[lev] = 현재 레벨 노드(0-based)

    def promote(node):
        if node["children"] is None:                   # leaf → group 승격
            node["children"] = []
            node["summary"] = node.get("values")
            node["values"] = None

    for r in range(data_start, ws.max_row + 1):
        labs = [_s(ws.cell(r, c).value) for c in range(1, LABELCOLS + 1)]
        vals = [_val(ws.cell(r, LABELCOLS + 1 + j)) for j in range(A)]
        kdeep = max((i + 1 for i, x in enumerate(labs) if x), default=0)
        if kdeep == 0:
            continue
        if labs[kdeep - 1].lower() in _SUMMARY_LABELS:  # 명시 '계'(위치무관) → 상위 그룹 summary
            tgt = path[kdeep - 2] if kdeep >= 2 else (roots[-1] if roots else None)
            if tgt is not None:
                promote(tgt)
                tgt["summary"] = vals
            continue
        # 한 행에 새 라벨이 여러 개일 수 있음(상위가 바뀌면 하위도 새로). 값은 가장 깊은 노드에.
        changed, deepest = False, None
        for lev in range(LABELCOLS):
            lab = labs[lev]
            if not lab:
                continue
            if changed or path[lev] is None or path[lev]["label"] != lab:
                node = {"label": lab, "children": None, "summary": None, "values": None}
                if lev >= 1 and path[lev - 1] is not None:
                    promote(path[lev - 1])
                    path[lev - 1]["children"].append(node)
                else:
                    roots.append(node)
                path[lev] = node
                for i in range(lev + 1, LABELCOLS):
                    path[i] = None
                changed, deepest = True, node
            else:
                deepest = path[lev]
        if deepest is not None and any(v is not None for v in vals):
            if deepest["children"] is None:
                deepest["values"] = vals
            else:
                deepest["summary"] = vals
    if not roots:
        raise ValueError("데이터 행을 찾지 못했습니다.")
    _autosum(roots, A)

    model = {"attributes": attrs, "nodes": roots, "level_labels": level_labels, "depth": LABELCOLS}
    for key, cell in (("title", "B1"), ("unit", "B2"), ("corner_label", "B3")):
        if _s(ws[cell].value):
            model[key] = _s(ws[cell].value)
    return model


def read_rowcol(ws):
    """라벨 열 개수를 자동 감지 → 2단계는 기존 items, 3+단계는 nodes 트리."""
    LABELCOLS = _detect_label_cols(ws)
    if LABELCOLS <= 2:
        return _read_rowcol_2level(ws)
    return _read_rowcol_ndepth(ws, LABELCOLS)


def _read_rowcol_2level(ws):
    attrs, c = [], 3
    while _s(ws.cell(5, c).value):
        attrs.append(_s(ws.cell(5, c).value))
        c += 1
    natt = len(attrs)
    if natt == 0:
        raise ValueError("행·열 시트 5행 C열부터 속성(매출/영업이익 …)을 1개 이상 적어주세요.")

    items, cur = [], None
    for r in range(6, ws.max_row + 1):
        a, b = _s(ws.cell(r, 1).value), _s(ws.cell(r, 2).value)
        vals = [_val(ws.cell(r, 3 + j)) for j in range(natt)]
        if a and (cur is None or a != cur["label"]):   # A 비었거나 직전과 같으면 같은 항목으로 이어붙임(병합/단일/반복 모두 지원)
            cur = {"label": a, "details": []}
            items.append(cur)
        if cur is None:
            continue
        if b.lower() in _SUMMARY_LABELS:
            cur["summary"] = vals
        elif b:
            cur["details"].append({"label": b, "values": vals})
    if not items:
        raise ValueError("행·열 시트 6행부터 데이터(항목/세부/값)를 채워주세요.")

    model = {"attributes": attrs, "items": items}
    for key, cell in (("title", "B1"), ("unit", "B2"), ("corner_label", "B3")):
        if _s(ws[cell].value):
            model[key] = _s(ws[cell].value)
    return model


# ── 읽기: 행+열(교차) 시트 → model ──────────────────────────────────────
def read_both(ws):
    cols, c = [], 3  # (열인덱스, 열그룹라벨, 열세부라벨)  ※ 병합/빈칸은 forward-fill
    last_cg = None
    while True:
        cg, cd = ws.cell(5, c).value, ws.cell(6, c).value
        if not _s(cg) and not _s(cd):
            break
        if _s(cg):
            last_cg = _s(cg)
        cols.append((c, last_cg, _s(cd) or None))
        c += 1
    if not cols:
        raise ValueError("행+열 시트 5~6행 C열부터 열그룹/열세부를 적어주세요.")

    col_groups, cg_map = [], {}
    for _, cg, cd in cols:
        if cg not in cg_map:
            cg_map[cg] = {"label": cg, "details": []}
            col_groups.append(cg_map[cg])
        if cd and cd not in cg_map[cg]["details"]:
            cg_map[cg]["details"].append(cd)

    row_groups, rg_map, data = [], {}, {}
    last_rg = None
    for r in range(7, ws.max_row + 1):
        a, b = _s(ws.cell(r, 1).value), _s(ws.cell(r, 2).value)
        if a:
            last_rg = a
        rg = last_rg
        if not rg or not b:
            continue
        if rg not in rg_map:
            rg_map[rg] = {"label": rg, "details": []}
            row_groups.append(rg_map[rg])
            data[rg] = {}
        rg_map[rg]["details"].append(b)
        data[rg][b] = {}
        for ci, cg, cd in cols:
            v = _val(ws.cell(r, ci))
            data[rg][b].setdefault(cg, {})[cd] = 0 if v is None else v
    if not row_groups:
        raise ValueError("행+열 시트 7행부터 행그룹/행세부/값을 채워주세요.")

    model = {"row_groups": row_groups, "col_groups": col_groups, "data": data}
    for key, cell in (("title", "B1"), ("unit", "B2"), ("corner_label", "B3")):
        if _s(ws[cell].value):
            model[key] = _s(ws[cell].value)
    return model


# ── Excel COM (DRM/보호·다중시트·비ooxml 대응) ──────────────────────────
# 참고: C:\codex\spoint (export_excel_to_txt.ps1 / TableFlowTxtExport.bas).
# openpyxl은 DRM(IRM/RMS)·암호화·구형 포맷을 못 연다. Excel COM은 '사용자에게 권한이
# 있으면' 그 파일을 열 수 있고, 평문 .xlsx 로 SaveAs(=복호화·정규화)하면 openpyxl이 읽는다.
_XL_OOXML = 51  # xlOpenXMLWorkbook


def _excel_app():
    import win32com.client as w  # Windows + Excel 설치 필요(지연 import)
    xl = w.DispatchEx("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        xl.EnableEvents = False
    except Exception:
        pass
    return xl


def normalize_via_com(path, password=None):
    """Excel COM으로 열어 평문 .xlsx 임시본으로 저장(복호화·정규화). 임시 경로 반환."""
    path = os.path.abspath(path)
    xl = _excel_app()
    try:
        if password:
            wb = xl.Workbooks.Open(Filename=path, ReadOnly=True, Password=password)
        else:
            wb = xl.Workbooks.Open(Filename=path, ReadOnly=True)
        fd, tmp = tempfile.mkstemp(prefix="dd_norm_", suffix=".xlsx")
        os.close(fd)
        os.remove(tmp)  # SaveAs가 새로 생성
        wb.SaveAs(tmp, FileFormat=_XL_OOXML)
        wb.Close(False)
        return tmp
    finally:
        try:
            xl.Quit()
        except Exception:
            pass


def extract_to_txt(path, out_dir, all_sheets=True, password=None):
    """spoint식: Excel COM으로 열어 시트별 TSV(UTF-8) 추출. 임의 파일·점검·진단용."""
    path = os.path.abspath(path)
    os.makedirs(out_dir, exist_ok=True)
    base = re.sub(r'[\\/:*?"<>|]', "_", os.path.splitext(os.path.basename(path))[0])
    xl = _excel_app()
    made = []
    try:
        if password:
            wb = xl.Workbooks.Open(Filename=path, ReadOnly=True, Password=password)
        else:
            wb = xl.Workbooks.Open(Filename=path, ReadOnly=True)
        sheets = list(wb.Worksheets) if all_sheets else [wb.Worksheets(1)]
        for ws in sheets:
            vals = ws.UsedRange.Value
            safe = re.sub(r'[\\/:*?"<>|]', "_", ws.Name)
            tp = os.path.join(out_dir, f"{base}__{safe}.txt" if all_sheets else f"{base}.txt")
            rows = vals if isinstance(vals, tuple) else ((vals,),)
            with open(tp, "w", encoding="utf-8-sig", newline="") as fh:
                for row in rows:
                    cells = row if isinstance(row, tuple) else (row,)
                    fh.write("\t".join(
                        "" if c is None else str(c).replace("\t", " ").replace("\r", " ").replace("\n", " ")
                        for c in cells) + "\n")
            made.append(tp)
        wb.Close(False)
        return made
    finally:
        try:
            xl.Quit()
        except Exception:
            pass


def _load_wb(path, com, password):
    """(workbook, 정리할_임시경로) 반환. com: True=강제COM, False=openpyxl만, 'auto'=실패시COM."""
    if com is True:
        tmp = normalize_via_com(path, password)
        return openpyxl.load_workbook(tmp, data_only=True), tmp
    try:
        return openpyxl.load_workbook(path, data_only=True), None
    except Exception:
        if com is False:
            raise
        tmp = normalize_via_com(path, password)   # DRM/암호화/구형 → Excel로 정규화
        return openpyxl.load_workbook(tmp, data_only=True), tmp


# ── 진입점: 파일 + orient → model ──────────────────────────────────────
def read_model(path, orient, com="auto", password=None):
    """입력 엑셀 → model dict.
    com='auto'(기본): openpyxl 먼저, DRM/암호화 등으로 실패하면 Excel COM으로 정규화 후 재시도.
    com=True: 처음부터 Excel COM(보호 파일 확실할 때). com=False: openpyxl만."""
    wb, tmp = _load_wb(path, com, password)
    try:
        if orient == "both":
            if SHEET_BOTH not in wb.sheetnames:
                raise ValueError(f"'{SHEET_BOTH}' 시트가 없습니다. 입력양식.xlsx 의 행+열 시트를 채워주세요.")
            return read_both(wb[SHEET_BOTH])
        if SHEET_ROWCOL not in wb.sheetnames:
            raise ValueError(f"'{SHEET_ROWCOL}' 시트가 없습니다. 입력양식.xlsx 의 행·열 시트를 채워주세요.")
        return read_rowcol(wb[SHEET_ROWCOL])
    finally:
        if tmp and os.path.exists(tmp):
            try:
                os.remove(tmp)
            except OSError:
                pass


# ── 클립보드/붙여넣기 (TSV) → model ─────────────────────────────────────
# 엑셀에서 표 영역을 복사하면 클립보드에 TSV(탭 구분, 행=개행)가 실린다.
# 붙여넣은 블록을 템플릿 데이터 시작(5행)에 올린 메모리 워크북으로 만들어 기존 파서를 재사용.
def _coerce(s):
    s = (s or "").strip()
    if not s:
        return None
    if s.endswith("%"):          # 퍼센트는 문자열 그대로 ("21%")
        return s
    t = s.replace(",", "")       # 천단위 구분 제거 후 숫자 시도
    try:
        fv = float(t)
        return int(fv) if fv.is_integer() else fv
    except ValueError:
        return s                 # 일반 텍스트


def _grid_to_wb(grid, orient):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_BOTH if orient == "both" else SHEET_ROWCOL
    for i, row in enumerate(grid):          # grid[0] → 시트 5행(헤더/열그룹), 데이터는 그 아래
        for j, cell in enumerate(row):
            v = _coerce(cell)
            if v is not None:
                ws.cell(5 + i, 1 + j, v)
    return wb


def read_model_from_text(text, orient):
    """엑셀에서 복사한 TSV(붙여넣기) → model.
    row/column: 첫 줄=헤더(항목·세부·속성…), 그 아래=데이터('계'=합계줄).
    both: 1줄=열그룹, 2줄=열세부, 3줄부터=행항목·행세부·교차값."""
    text = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    grid = [ln.split("\t") for ln in text.split("\n")]
    grid = [r for r in grid if any((c or "").strip() for c in r)]   # 완전 빈 줄 제거
    if not grid:
        raise ValueError("붙여넣은 내용이 비어 있습니다. 엑셀에서 표 영역을 복사한 뒤 붙여넣어 주세요.")
    ws = _grid_to_wb(grid, orient).active
    return read_both(ws) if orient == "both" else read_rowcol(ws)


# ── 입력양식 생성 (예시 데이터 미리채움 + 스타일) ────────────────────────
_GRAY = PatternFill("solid", fgColor="E7E6E6")
_HEAD = PatternFill("solid", fgColor="D6E0F2")
_BOLD = Font(bold=True)
_NOTE = Font(italic=True, color="808080", size=9)
_CEN = Alignment(horizontal="center", vertical="center")
_thin = Side(style="thin", color="BFBFBF")
_BORD = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _meta(ws, title, unit, corner):
    for i, (lab, val) in enumerate((("제목", title), ("단위", unit), ("코너라벨", corner)), start=1):
        ws.cell(i, 1, lab).fill = _GRAY
        ws.cell(i, 1).font = _BOLD
        ws.cell(i, 2, val)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16


def _build_rowcol_sheet(ws):
    _meta(ws, "2024년 사업부별 실적", "단위: 억원 / %", "사업부")
    ws.cell(4, 1, "↓ 5행은 헤더(항목·세부·속성), 6행부터 데이터. '계' 행=항목 합계줄.").font = _NOTE
    head = ["항목", "세부", "매출", "영업이익", "점유율"]
    for j, h in enumerate(head, start=1):
        cell = ws.cell(5, j, h)
        cell.font, cell.fill, cell.alignment, cell.border = _BOLD, _HEAD, _CEN, _BORD
    rows = [
        ("A사업부", "계", 1200, 180, "21%"), ("", "Q1", 280, 38, "20%"),
        ("", "Q2", 300, 42, "21%"), ("", "Q3", 300, 48, "22%"), ("", "Q4", 320, 52, "21%"),
        ("B사업부", "계", 800, 90, "12%"), ("", "Q1", 180, 18, "11%"),
        ("", "Q2", 200, 22, "12%"), ("", "Q3", 210, 24, "12%"), ("", "Q4", 210, 26, "13%"),
        ("C사업부", "계", 2100, 520, "33%"), ("", "Q1", 480, 110, "31%"),
        ("", "Q2", 510, 125, "32%"), ("", "Q3", 540, 135, "34%"), ("", "Q4", 570, 150, "35%"),
    ]
    for i, row in enumerate(rows, start=6):
        for j, v in enumerate(row, start=1):
            c = ws.cell(i, j, v)
            c.border = _BORD
            if j >= 3 and isinstance(v, str) and v.endswith("%"):
                c.number_format = "@"  # 퍼센트는 텍스트로 보존
    for col, w in (("C", 11), ("D", 11), ("E", 11)):
        ws.column_dimensions[col].width = w


def _build_both_sheet(ws):
    _meta(ws, "사업부 × 연도 매출", "단위: 억원", "사업부 ＼ 연도")
    ws.cell(4, 1, "↓ 5행=열그룹, 6행=열세부 / 7행부터 행그룹·행세부·교차값. 계/총계는 자동 합산.").font = _NOTE
    # 열 헤더 (C열부터): 5행 연도(병합), 6행 반기
    ws.cell(6, 1, "행항목").font = _BOLD
    ws.cell(6, 2, "행세부").font = _BOLD
    ws.cell(6, 1).fill = ws.cell(6, 2).fill = _HEAD
    year_cols = [("2023", 3), ("2024", 5)]
    for yr, c0 in year_cols:
        ws.merge_cells(start_row=5, start_column=c0, end_row=5, end_column=c0 + 1)
        yc = ws.cell(5, c0, yr)
        yc.font, yc.fill, yc.alignment = _BOLD, _HEAD, _CEN
        for off, half in enumerate(("상반기", "하반기")):
            hc = ws.cell(6, c0 + off, half)
            hc.font, hc.fill, hc.alignment, hc.border = _BOLD, _HEAD, _CEN, _BORD
    data = [
        ("A사업부", "국내", 700, 750, 780, 820), ("", "해외", 450, 500, 470, 530),
        ("B사업부", "국내", 300, 320, 310, 330), ("", "해외", 250, 260, 270, 280),
    ]
    for i, row in enumerate(data, start=7):
        for j, v in enumerate(row, start=1):
            c = ws.cell(i, j, v)
            c.border = _BORD
            if j <= 2:
                c.alignment = Alignment(horizontal="left")
    for col, w in (("A", 12), ("B", 10), ("C", 9), ("D", 9), ("E", 9), ("F", 9)):
        ws.column_dimensions[col].width = w


def make_template(path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = SHEET_ROWCOL
    _build_rowcol_sheet(ws1)
    _build_both_sheet(wb.create_sheet(SHEET_BOTH))
    wb.save(path)
    return path


# ── 자체 테스트 ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    import os, sys, tempfile, json
    sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), os.pardir))
    import drilldown_table as G

    tmp = os.path.join(tempfile.gettempdir(), "_dd_template_test.xlsx")
    make_template(tmp)
    print("template ->", tmp)
    m_rc = read_model(tmp, "row")
    m_2d = read_model(tmp, "both")
    print("rowcol items:", [it["label"] for it in m_rc["items"]],
          "| attrs:", m_rc["attributes"])
    print("both rows:", [g["label"] for g in m_2d["row_groups"]],
          "| cols:", [(g["label"], g["details"]) for g in m_2d["col_groups"]])
    # 모델 동일성 확인
    assert m_rc == G.SAMPLE, "rowcol 모델이 엔진 SAMPLE 과 다릅니다:\n" + json.dumps(m_rc, ensure_ascii=False, indent=1)
    assert m_2d == G.SAMPLE_2D, "both 모델이 엔진 SAMPLE_2D 와 다릅니다:\n" + json.dumps(m_2d, ensure_ascii=False, indent=1)
    print("OK: 입력양식 round-trip 이 엔진 SAMPLE/SAMPLE_2D 와 100% 일치")
